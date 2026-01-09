"""
Excel Export Service for generating bill reports with date filtering
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from typing import List, Optional
import os
import logging
from models import Bill

logger = logging.getLogger(__name__)

EXPORTS_FOLDER = "/app/exports"


class ExportService:
    """
    Service for exporting bills to Excel with date range filtering
    """
    
    def __init__(self):
        # Ensure exports folder exists
        os.makedirs(EXPORTS_FOLDER, exist_ok=True)
    
    def generate_excel(
        self,
        bills: List[Bill],
        user_id: str,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None
    ) -> str:
        """
        Generate Excel file from bills with optional date filtering
        
        Args:
            bills: List of Bill objects
            user_id: User identifier
            start_date: Optional start date filter
            end_date: Optional end date filter
            
        Returns:
            Path to generated Excel file
        """
        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Bills"
            
            # Style definitions
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            headers = [
                "Date", "Shop Name", "Shop Type", "Location",
                "Total Price", "Currency", "Tax Amount",
                "Description", "Status", "Items Count"
            ]
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Filter bills by date if provided
            filtered_bills = bills
            if start_date or end_date:
                filtered_bills = self._filter_bills_by_date(bills, start_date, end_date)
            
            # Add bill data
            for row_num, bill in enumerate(filtered_bills, 2):
                items_count = len(bill.menu) if bill.menu else 0
                
                data = [
                    bill.created_at.strftime("%Y-%m-%d %H:%M") if bill.created_at else "",
                    bill.shop_name or "N/A",
                    bill.shop_type or "N/A",
                    bill.location or "N/A",
                    float(bill.total_price) if bill.total_price else 0,
                    bill.currency or "USD",
                    float(bill.tax_amount) if bill.tax_amount else 0,
                    bill.description or "",
                    bill.status or "processed",
                    items_count
                ]
                
                for col_num, value in enumerate(data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    
                    # Align numbers to right
                    if col_num in [5, 7, 10]:
                        cell.alignment = Alignment(horizontal='right')
            
            # Create detailed menu sheet
            ws_menu = wb.create_sheet(title="Menu Items")
            self._create_menu_sheet(ws_menu, filtered_bills)
            
            # Add summary sheet
            ws_summary = wb.create_sheet(title="Summary")
            self._create_summary_sheet(ws_summary, filtered_bills)
            
            # Adjust column widths
            column_widths = [18, 20, 15, 25, 12, 10, 12, 30, 12, 12]
            for col_num, width in enumerate(column_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
            
            # Generate filename
            date_suffix = ""
            if start_date and end_date:
                date_suffix = f"_{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d')}"
            
            filename = f"bills_{user_id}{date_suffix}.xlsx"
            filepath = os.path.join(EXPORTS_FOLDER, filename)
            
            # Save workbook
            wb.save(filepath)
            
            logger.info(f"Excel export generated: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error generating Excel: {str(e)}")
            raise
    
    def _filter_bills_by_date(
        self,
        bills: List[Bill],
        start_date: Optional[datetime],
        end_date: Optional[datetime]
    ) -> List[Bill]:
        """
        Filter bills by date range
        """
        filtered = []
        for bill in bills:
            if not bill.created_at:
                continue
            
            if start_date and bill.created_at < start_date:
                continue
            
            if end_date and bill.created_at > end_date:
                continue
            
            filtered.append(bill)
        
        return filtered
    
    def _create_menu_sheet(self, ws, bills: List[Bill]):
        """
        Create detailed menu items sheet
        """
        # Headers
        headers = ["Date", "Shop Name", "Item Name", "Quantity", "Price"]
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        row_num = 2
        for bill in bills:
            if not bill.menu:
                continue
            
            for item in bill.menu:
                ws.cell(row=row_num, column=1).value = bill.created_at.strftime("%Y-%m-%d") if bill.created_at else ""
                ws.cell(row=row_num, column=2).value = bill.shop_name or "N/A"
                ws.cell(row=row_num, column=3).value = item.get("item", "")
                ws.cell(row=row_num, column=4).value = item.get("quantity", 0)
                ws.cell(row=row_num, column=5).value = item.get("price", 0)
                row_num += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
    
    def _create_summary_sheet(self, ws, bills: List[Bill]):
        """
        Create summary statistics sheet
        """
        # Title
        ws['A1'] = "Bills Summary Report"
        ws['A1'].font = Font(bold=True, size=16)
        
        # Statistics
        total_bills = len(bills)
        total_amount = sum(float(bill.total_price) for bill in bills if bill.total_price)
        total_tax = sum(float(bill.tax_amount) for bill in bills if bill.tax_amount)
        
        # Shop types breakdown
        shop_types = {}
        for bill in bills:
            shop_type = bill.shop_type or "Unknown"
            shop_types[shop_type] = shop_types.get(shop_type, 0) + 1
        
        # Write statistics
        ws['A3'] = "Total Bills:"
        ws['B3'] = total_bills
        ws['A4'] = "Total Amount:"
        ws['B4'] = round(total_amount, 2)
        ws['A5'] = "Total Tax:"
        ws['B5'] = round(total_tax, 2)
        
        # Shop types breakdown
        ws['A7'] = "Bills by Shop Type"
        ws['A7'].font = Font(bold=True)
        
        row = 8
        for shop_type, count in shop_types.items():
            ws.cell(row=row, column=1).value = shop_type
            ws.cell(row=row, column=2).value = count
            row += 1


# Singleton instance
export_service = ExportService()
